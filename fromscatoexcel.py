# -*- coding: utf-8 -*-
import os
import openpyxl
import numpy as np


def readsca():
    path_part1 = "D:/data/3 DDA/#18 锥形硅纳米线/文章2补/onesection-air-20-120-1um-"
    angle = 0, 30, 60, 90, 120, 150, 180

    for i in range(0, 7):
        path_part2 = os.path.join(str(angle[i]))
        path1 = os.path.join(path_part1 + path_part2)
        path2 = os.path.join(path_part1 + path_part2 + "-2")
        pol1 = np.zeros((51, 7), dtype="float64")
        pol2 = np.zeros((51, 7), dtype="float64")
        lnum2 = 0  # pol中行号

        for root, dirs, filenames in os.walk(path1):
            count = 0
            for filename in filenames:
                if os.path.splitext(filename)[1] == '.sca':
                    count += 1

            for count1 in range(0, 10):
                filename = os.path.join("w00" + str(count1) + "r000k000" + ".sca")
                f = open(os.path.join(path1, filename), "r")  # 打开文件

                lnum = 0  # txt中行号

                fd = f.readlines()  # 读取全部内容
                for line in fd:
                    lnum += 1
                    if lnum == 35:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol1[lnum2] = line[2:9]  # 数组赋值

                    if lnum == 36:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol2[lnum2] = line[2:9]  # 数组赋值
                        lnum2 += 1

                f.close()

            for count2 in range(10, count):
                filename = os.path.join("w0" + str(count2) + "r000k000" + ".sca")
                f = open(os.path.join(path1, filename), "r")  # 打开文件

                lnum = 0  # txt中行号

                fd = f.readlines()  # 读取全部内容
                for line in fd:
                    lnum += 1
                    if lnum == 35:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol1[lnum2] = line[2:9]  # 数组赋值

                    if lnum == 36:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol2[lnum2] = line[2:9]  # 数组赋值
                        lnum2 += 1

                f.close()

        for root, dirs, filenames2 in os.walk(path2):
            count = 0
            for filename in filenames2:
                if os.path.splitext(filename)[1] == ".sca":
                    count += 1

            for count1 in range(0, 10):
                filename = os.path.join("w00" + str(count1) + "r000k000" + ".sca")
                f = open(os.path.join(path2, filename), "r")  # 打开文件

                lnum = 0  # txt中行号

                fd = f.readlines()  # 读取全部内容
                for line in fd:
                    lnum += 1
                    if lnum == 35:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol1[lnum2] = line[2:9]  # 数组赋值

                    if lnum == 36:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol2[lnum2] = line[2:9]  # 数组赋值
                        lnum2 += 1

                f.close()

            for count3 in range(10, count):
                filename = os.path.join("w0" + str(count3) + "r000k000" + ".sca")
                f = open(os.path.join(path2, filename), "r")  # 打开文件

                lnum = 0  # txt中行号

                fd = f.readlines()  # 读取全部内容
                for line in fd:
                    lnum += 1
                    if lnum == 35:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol1[lnum2] = line[2:9]  # 数组赋值

                    if lnum == 36:
                        line = line.replace('  ', ' ').replace('\n', '').replace(',', ' ').split(' ')
                        pol2[lnum2] = line[2:9]  # 数组赋值
                        lnum2 += 1

                f.close()

        file_2007 = os.path.join("D:/data/3 DDA/#18 锥形硅纳米线/文章2补", str(angle[i]) + ".xlsx")
        write07excel(file_2007, pol1, pol2)


if __name__ == "__readsca__":
    readsca()


def write07excel(path, pol1, pol2):
    wb = openpyxl.Workbook()
    ws1 = wb.create_sheet("pol1", 0)
    ws2 = wb.create_sheet("pol2", 1)
    for i in range(0, 51):
        for j in range(0, 7):
            ws1.cell(row=i + 1, column=j + 1, value=pol1[i, j])
            ws2.cell(row=i + 1, column=j + 1, value=pol2[i, j])
    wb.save(path)
